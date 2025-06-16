import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import tempfile

# Titre de l'application
st.title("📦 Générateur de Reporting Ventes SIM")

# Uploader du fichier Excel brut
uploaded_file = st.file_uploader("📁 Importer le fichier Excel brut (hebdomadaire)", type=["xlsx"])

if uploaded_file:
    # Charger toutes les feuilles sans les lire entièrement
    xls = pd.ExcelFile(uploaded_file)
    
    # Afficher les noms de feuilles disponibles
    sheet_names = xls.sheet_names
    selected_sheet = st.selectbox("🗂️ Choisir la feuille à exploiter :", options=sheet_names)
    
    # Lire uniquement la feuille sélectionnée
    df = pd.read_excel(uploaded_file, sheet_name=selected_sheet)
    
    st.success(f"✅ Feuille '{selected_sheet}' chargée avec succès !")
    st.dataframe(df.head())


    # Nettoyage / Préparation
    df = df.rename(columns={'MSISDN': 'TOTAL_SIM'})
    df = df.rename(columns={'ACCUEIL_VENDEUR': 'PVT'})
    df = df.rename(columns={'LOGIN_VENDEUR': 'LOGIN'})
    df = df.rename(columns={'AGENCE_VENDEUR': 'DRV'})

    # -------- Résumé par VTO --------
    
    df_summary = df.groupby(['DRV', 'PVT', 'PRENOM_VENDEUR', 'NOM_VENDEUR', 'LOGIN']).agg({
    'TOTAL_SIM': 'count'}).reset_index()  
    
    # Trier les données pour regrouper visuellement
    df_summary = df_summary.sort_values(['DRV', 'PVT'])

    # Pour masquer les répétitions (laisser vide sauf première occurrence)
    #df_summary['DRV'] = df_summary['DRV'].mask(df_summary['DRV'].duplicated())
    #df_summary['PVT'] = df_summary['PVT'].mask(df_summary['PVT'].duplicated())

    df_summary["DRV"] = df_summary["DRV"].replace({ 
    "DV-DRV2_DIRECTION REGIONALE DES VENTES DAKAR 2": "DR2",
    "DV-DRVS_DIRECTION REGIONALE DES VENTES SUD": "DR SUD",
    "DV-DRVSE_DIRECTION REGIONALE DES VENTES SUD-EST": "SUD EST",
    "DV-DRVN_DIRECTION REGIONALE DES VENTES NORD": "DR NORD",
    "DV-DRVC_DIRECTION REGIONALE DES VENTES CENTRE": "DR CENTRE",
    "DV-DRVE_DIRECTION REGIONALE DES VENTES EST": "DR EST"
     })


    # Pour masquer les répétitions (laisser vide sauf première occurrence)
    df_summaryy = df_summary.copy()
    df_summaryy['DRV'] = df_summaryy['DRV'].mask(df_summaryy['DRV'].duplicated())
    df_summaryy['PVT'] = df_summaryy['PVT'].mask(df_summaryy['PVT'].duplicated())


    # -------- Ventes par PVT (si dispo) --------
    df_summary2 = df.groupby(['DRV', 'PVT']).agg({
    'TOTAL_SIM': 'count'}).reset_index()

    #------------------------------------------------------------------------------------------------
    #Pour fusionner les lignes vides

    # 1. Créer un fichier Excel temporaire avec pandas
    temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    with pd.ExcelWriter(temp_file.name, engine='openpyxl') as writer:
        df_summaryy.to_excel(writer, sheet_name='Résumé Ventes', index=False)
        df_summary2.to_excel(writer, sheet_name='Ventes Par PVT', index=False)

    # 2. Charger avec openpyxl pour appliquer la fusion
    wb = load_workbook(temp_file.name)
    ws = wb["Résumé Ventes"]

    

    # 3. Sauvegarde dans un buffer pour Streamlit
    final_buffer = BytesIO()
    wb.save(final_buffer)
    final_buffer.seek(0)
        

     #----------------------------------------------------------------------------------------------------

    # Télécharger le fichier généré
    st.success("✅ Fichier généré avec succès !")
    st.download_button(
        label="📥 Télécharger le fichier Excel",
        data=final_buffer,
        file_name="Weekly Reporting.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )




    import pandas as pd
    from datetime import datetime
    import re
    from io import BytesIO
    import os

    fichier_historique = r"C:\Users\hp\Downloads\Dossier LOUMA\historique_ventes_.xlsx"

    # 1. Préparer df_summary
    df_history = df_summary.copy()

    # 2. Identifier la dernière semaine existante
    if os.path.exists(fichier_historique):
        df_old = pd.read_excel(fichier_historique)

        # Extraire toutes les colonnes de type SEMxx
        sem_cols = [col for col in df_old.columns if re.match(r"SEM\d+", str(col))]
        if sem_cols:
            derniers_num_semaines = [int(col[3:]) for col in sem_cols]
            max_semaine = max(derniers_num_semaines)
            nouvelle_semaine = max_semaine + 1
        else:
            nouvelle_semaine = 1

        nom_col_semaine = f"SEM{nouvelle_semaine}"
        st.info(f"🕓 Semaine détectée : {nom_col_semaine}")

        #
        #yallah pitié
        def create_key(df):
         return (
            df_summary['DRV'].astype(str).str.strip().str.upper() + "|" +
            df_summary['PVT'].astype(str).str.strip().str.upper() + "|" +
            df_summary['PRENOM_VENDEUR'].astype(str).str.strip().str.upper() + "|" +
            df_summary['NOM_VENDEUR'].astype(str).str.strip().str.upper()
        )

        df_old["KEY"] = create_key(df_old)

        # Préparer df_summary de la semaine courante
        
        df_history["KEY"] = create_key(df_history)
        df_history = df_history[["KEY", "TOTAL_SIM"]].copy()
        
        #
        # Renommer la colonne dans df_summary
        df_history = df_history.rename(columns={'TOTAL_SIM': nom_col_semaine})
        #
        #df_history_clean = df_history[['DRV', 'PVT', 'PRENOM_VENDEUR', 'NOM_VENDEUR', nom_col_semaine]]
        df_history_clean = df_history[["KEY", nom_col_semaine]]
        
        # Renommer la colonne dans df_summary
        #df_history_clean = df_history_clean.rename(columns={'TOTAL_SIM': nom_col_semaine})

        

        # Fusionner avec l'existant
        #df_merged = pd.merge(df_old, df_history_clean, on=['DRV', 'PVT', 'PRENOM_VENDEUR', 'NOM_VENDEUR'], how='outer')

        df_merged = pd.merge(df_old, df_history_clean, on="KEY", how="left")
        df_merged[nom_col_semaine] = df_merged[nom_col_semaine].fillna(0)

        df_merged = df_merged.drop(columns=["KEY"])



    else:
        nom_col_semaine = "SEM1"
        st.info(f"📁 Fichier historique non trouvé : création avec {nom_col_semaine}")
        df_history = df_history.rename(columns={'TOTAL_SIM': nom_col_semaine})
        df_merged = df_history.copy()

    # 3. Sauvegarder le fichier mis à jour
    df_merged.to_excel(fichier_historique, index=False)
    st.success(f"✅ Données ajoutées à l’historique (colonne : {nom_col_semaine})")

    # 4. Ajouter un bouton pour télécharger le fichier historique
    with open(fichier_historique, "rb") as file:
        st.download_button(
            label="📥 Télécharger le fichier historique",
            data=file,
            file_name="historique_ventes.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
